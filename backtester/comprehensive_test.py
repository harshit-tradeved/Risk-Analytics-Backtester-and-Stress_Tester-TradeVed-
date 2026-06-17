"""
Comprehensive Feature Test & Report Generator
=============================================
Exhaustively exercises EVERY feature added since the original GRID/DCA/PLA
backtester + stress tester, then writes Excel reports in the same style as the
earlier `qa_reports/` workbooks.

Coverage
--------
1. STRATEGY MATRIX   — all 54 registered strategies × every asset class
                       (crypto / US stocks / Indian equity) × multiple
                       timeframes (1d / 4h / 1h). Mirrors main.py's run path
                       (auto GRID bounds, Indian cost auto-detect, lot sizes,
                       metrics, regime classification).
2. STRESS MATRIX     — representative strategies × all 17 stress scenarios ×
                       3 severities on crypto + Indian assets.
3. AI FORECAST       — block-bootstrap forward-test, crisis-overlay, and
                       paper-trading paths on selected strategy/asset combos
                       (the Kronos-ready pipeline; GPU not required).
4. INDICATOR ENGINE  — every indicator in INDICATOR_CATALOG computes cleanly
                       with stable output columns.
5. EXECUTION SUMMARY — pass/fail counts, coverage, environment, defect log.

Usage:
    python comprehensive_test.py                 # full run
    python comprehensive_test.py --quick         # smaller universe (smoke)
    python comprehensive_test.py --skip-ai       # skip forecast section

Output:
    comprehensive_reports/
        Strategy_Backtest_Matrix.xlsx
        Stress_Test_Matrix.xlsx
        AI_Forecast_Report.xlsx
        Indicator_Validation.xlsx
        Execution_Summary.xlsx
        Defect_Log.xlsx
        raw_results.csv
"""
from __future__ import annotations

import argparse
import logging
import sys
import time
import traceback
import warnings
from collections import Counter
from datetime import datetime
from pathlib import Path

import numpy as np
import pandas as pd

# ── Quiet noise ──────────────────────────────────────────────────────────────
warnings.filterwarnings("ignore")
logging.disable(logging.WARNING)
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

HERE = Path(__file__).parent
sys.path.insert(0, str(HERE))

from data.fetcher import DataFetcher
from data.validator import DataValidator
from data.indian_assets import is_indian, get_lot_size
from engine.simulator import TradeSimulator
from engine.metrics import calculate_metrics
from engine.regimes import classify_regimes
from engine import stress as stress_engine
from engine import forecast as forecast_engine
from engine.indicators import INDICATOR_CATALOG, compute
from strategies import STRATEGY_REGISTRY

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────────────────────
# Universe configuration
# ─────────────────────────────────────────────────────────────────────────────

# (symbol, source, asset_class, currency)
CRYPTO = [("BTC/USDT", "binance", "Crypto", "USD"),
          ("ETH/USDT", "binance", "Crypto", "USD"),
          ("SOL/USDT", "binance", "Crypto", "USD"),
          ("BNB/USDT", "binance", "Crypto", "USD")]
US = [("AAPL", "yfinance", "US Stock", "USD"),
      ("TSLA", "yfinance", "US Stock", "USD"),
      ("MSFT", "yfinance", "US Stock", "USD"),
      ("SPY",  "yfinance", "US Stock", "USD")]
INDIA = [("RELIANCE", "nse", "Indian Equity", "INR"),
         ("TCS",      "nse", "Indian Equity", "INR"),
         ("INFY",     "nse", "Indian Equity", "INR"),
         ("HDFCBANK", "nse", "Indian Equity", "INR")]

QUICK_CRYPTO = CRYPTO[:1]
QUICK_US = US[:1]
QUICK_INDIA = INDIA[:1]

# Timeframe windows (interval → (start, end)). 1h/4h windows kept short to keep
# candle counts and fetch payloads sane; 1d covers a full 2-year cycle.
TF_WINDOWS = {
    "1d":  (datetime(2022, 6, 1), datetime(2024, 6, 1)),
    "4h":  (datetime(2023, 9, 1), datetime(2024, 6, 1)),
    "1h":  (datetime(2024, 2, 1), datetime(2024, 6, 1)),
    "15m": (datetime(2024, 4, 1), datetime(2024, 6, 1)),
}

ALL_STRATEGIES = list(STRATEGY_REGISTRY.keys())
CAPITAL_USD = 10_000.0
CAPITAL_INR = 1_000_000.0

# Strategies used for the heavier stress / AI sweeps (one per family).
REP_STRATEGIES = ["DCA", "GRID", "PLA", "RSI", "MACD", "SUPERTREND",
                  "BOLLINGER", "MACROSS", "RSI_MACD", "CUSTOM"]

STRESS_SEVERITIES = [0.5, 1.0, 1.5]


# ─────────────────────────────────────────────────────────────────────────────
# Core backtest (mirrors main.py:run_backtest)
# ─────────────────────────────────────────────────────────────────────────────

def _round_nice(x: float, mode: str) -> float:
    if x <= 0:
        return 0.0
    mag = 10 ** np.floor(np.log10(abs(x)))
    return float(np.floor(x / mag) * mag) if mode == "floor" else float(np.ceil(x / mag) * mag)


def _ensure_grid_bounds(params: dict, df: pd.DataFrame) -> None:
    lo = float(params.get("lower_bound", 0) or 0)
    hi = float(params.get("upper_bound", 0) or 0)
    prices = df["close"].astype(float)
    lo_raw, hi_raw = float(prices.min()), float(prices.max())
    if lo >= hi or lo > hi_raw or hi < lo_raw:
        pad = (hi_raw - lo_raw) * 0.10
        params["lower_bound"] = _round_nice(max(lo_raw * 0.5, lo_raw - pad), "floor")
        params["upper_bound"] = _round_nice(hi_raw + pad, "ceil")


def regime_mix(df: pd.DataFrame) -> str:
    labels = classify_regimes(df)
    if not labels:
        return "n/a"
    c = Counter(labels)
    n = len(labels)
    return " ".join(f"{k[:4]}:{round(100 * v / n)}%" for k, v in
                    sorted(c.items(), key=lambda kv: -kv[1]))


def run_backtest(df, symbol, source, strat_name, asset_class, currency):
    """One backtest, returning a flat result row. Errors captured, not raised."""
    capital = CAPITAL_INR if currency == "INR" else CAPITAL_USD
    cls = STRATEGY_REGISTRY[strat_name]
    params = dict(cls.default_params())
    if strat_name == "GRID":
        _ensure_grid_bounds(params, df)

    use_indian = source in ("nse", "bse") or is_indian(symbol)
    market_type = "equity_delivery" if use_indian else "crypto"
    lot_sz = 1  # equity delivery → no lot enforcement

    inst = cls(**params)
    sigs = inst.generate_signals(df.copy())
    n_buy = int((sigs["signal"] == "BUY").sum())
    n_sell = int((sigs["signal"] == "SELL").sum())

    sim = TradeSimulator(
        symbol=symbol, capital=capital, use_indian_costs=use_indian,
        market_type=market_type, lot_size=lot_sz,
    )
    out = sim.run(sigs)
    m = calculate_metrics(out["trades"], out["equity_curve"],
                          out["timestamps"], capital)
    return {
        "asset_class": asset_class, "symbol": symbol, "source": source,
        "currency": currency, "strategy": strat_name,
        "category": cls.category() if hasattr(cls, "category") else "classic",
        "candles": len(df), "buy_signals": n_buy, "sell_signals": n_sell,
        "num_trades": m["num_trades"],
        "return_pct": round(m["total_return_pct"], 2),
        "ann_return_pct": round(m["annualised_return_pct"], 2),
        "sharpe": round(m["sharpe_ratio"], 3),
        "sortino": round(m["sortino_ratio"], 3),
        "calmar": round(m["calmar_ratio"], 3),
        "max_dd_pct": round(m["max_drawdown_pct"], 2),
        "win_rate": round(m["win_rate"], 1),
        "profit_factor": round(m["profit_factor"], 2),
        "final_equity": round(m["final_equity"], 2),
        "regime_mix": regime_mix(df),
        "status": "PASS", "error": "",
    }


def error_row(symbol, source, strat_name, asset_class, currency, exc):
    return {
        "asset_class": asset_class, "symbol": symbol, "source": source,
        "currency": currency, "strategy": strat_name, "category": "",
        "candles": 0, "buy_signals": 0, "sell_signals": 0, "num_trades": 0,
        "return_pct": None, "ann_return_pct": None, "sharpe": None,
        "sortino": None, "calmar": None, "max_dd_pct": None, "win_rate": None,
        "profit_factor": None, "final_equity": None, "regime_mix": "",
        "status": "ERROR", "error": f"{type(exc).__name__}: {exc}",
    }


# ─────────────────────────────────────────────────────────────────────────────
# Section 1 — Strategy × Asset × Timeframe matrix
# ─────────────────────────────────────────────────────────────────────────────

def section_strategy_matrix(fetcher, validator, universe, timeframes, defects):
    print("\n=== SECTION 1: Strategy × Asset × Timeframe matrix ===")
    rows = []
    # Build the (symbol, source, interval) jobs; fetch each df once, run all strats.
    for interval in timeframes:
        start, end = TF_WINDOWS[interval]
        for symbol, source, asset_class, currency in universe:
            # crypto handles all intervals; stocks/india reliably only 1d here
            if interval != "1d" and source != "binance":
                continue
            tag = f"{symbol} [{source}] {interval}"
            try:
                df = fetcher.fetch(symbol, start, end, source=source, interval=interval)
            except Exception as exc:
                print(f"  FETCH-FAIL {tag}: {type(exc).__name__}: {str(exc)[:60]}")
                defects.append(("FETCH", tag, f"{type(exc).__name__}: {exc}"))
                continue
            val = validator.validate(df)
            if not val.passed:
                print(f"  DATA-QUAL {tag}: {val.quality_score:.0f}/100")
                defects.append(("DATA_QUALITY", tag, f"score {val.quality_score:.0f}; {val.issues}"))
                continue

            ok = err = 0
            t0 = time.time()
            for strat in ALL_STRATEGIES:
                try:
                    r = run_backtest(df, symbol, source, strat, asset_class, currency)
                    r["interval"] = interval
                    rows.append(r)
                    ok += 1
                except Exception as exc:
                    er = error_row(symbol, source, strat, asset_class, currency, exc)
                    er["interval"] = interval
                    rows.append(er)
                    defects.append(("STRATEGY_RUN", f"{strat} on {tag}",
                                    f"{type(exc).__name__}: {exc}"))
                    err += 1
            print(f"  {tag:32} candles={len(df):5} strat ok={ok:3} err={err:2} "
                  f"({time.time()-t0:.1f}s)")
    return rows


# ─────────────────────────────────────────────────────────────────────────────
# Section 2 — Stress matrix (all 17 scenarios)
# ─────────────────────────────────────────────────────────────────────────────

def section_stress_matrix(fetcher, validator, defects, quick=False):
    print("\n=== SECTION 2: Stress matrix (17 scenarios) ===")
    rows = []
    assets = ([("BTC/USDT", "binance", "USD"), ("RELIANCE", "nse", "INR")]
              if not quick else [("BTC/USDT", "binance", "USD")])
    strategies = ["DCA", "PLA", "RSI"] if not quick else ["DCA"]
    scenarios = list(stress_engine.SCENARIO_PRESETS.keys())
    start, end = TF_WINDOWS["1d"]

    for symbol, source, currency in assets:
        capital = CAPITAL_INR if currency == "INR" else CAPITAL_USD
        try:
            df = fetcher.fetch(symbol, start, end, source=source, interval="1d")
        except Exception as exc:
            defects.append(("STRESS_FETCH", symbol, str(exc)))
            continue
        use_indian = source in ("nse", "bse") or is_indian(symbol)
        sim_kwargs = dict(symbol=symbol, use_indian_costs=use_indian,
                          market_type="equity_delivery" if use_indian else "crypto",
                          lot_size=1)
        for strat in strategies:
            cls = STRATEGY_REGISTRY[strat]
            params = dict(cls.default_params())
            if strat == "GRID":
                _ensure_grid_bounds(params, df)
            for scen in scenarios:
                scen_obj = stress_engine.SCENARIO_PRESETS[scen]
                for sev in STRESS_SEVERITIES:
                    try:
                        res = stress_engine.run_stress_backtest(
                            df=df, strategy_cls=cls, strategy_params=params,
                            sim_kwargs=sim_kwargs, capital=capital,
                            scenario=scen_obj, severity=sev,
                            monte_carlo_runs=30, seed=42,
                        )
                        base = res["baseline"]; st = res["stressed"]; mc = res["monte_carlo"]
                        delta = st["return_pct"] - base["total_return_pct"]
                        rows.append({
                            "symbol": symbol, "currency": currency, "strategy": strat,
                            "scenario": scen, "severity": sev,
                            "baseline_return_pct": round(base["total_return_pct"], 2),
                            "stressed_return_pct": round(st["return_pct"], 2),
                            "delta_pct": round(delta, 2),
                            "stressed_sharpe": round(st.get("sharpe", 0), 3),
                            "stressed_max_dd_pct": round(st.get("max_dd_pct", 0), 2),
                            "mc_return_p50": round(mc.get("return_pct", {}).get("p50", 0), 2),
                            "mc_return_p5": round(mc.get("return_pct", {}).get("p5", 0), 2),
                            "verdict": "SURVIVED" if delta >= -2 else
                                       ("DEGRADED" if delta >= -10 else "BROKEN"),
                            "status": "PASS",
                        })
                    except Exception as exc:
                        defects.append(("STRESS_RUN", f"{strat}/{scen}/{sev} on {symbol}",
                                        f"{type(exc).__name__}: {exc}"))
                        rows.append({
                            "symbol": symbol, "currency": currency, "strategy": strat,
                            "scenario": scen, "severity": sev, "verdict": "ERROR",
                            "status": "ERROR", "error": f"{type(exc).__name__}: {exc}",
                        })
            print(f"  {symbol:10} {strat:5} — {len(scenarios)} scenarios × "
                  f"{len(STRESS_SEVERITIES)} sev done")
    return rows


# ─────────────────────────────────────────────────────────────────────────────
# Section 3 — AI forecast (forward / crisis / paper)
# ─────────────────────────────────────────────────────────────────────────────

def _aggregate_paths(df_ctx, paths, cls, params, sim_kwargs, capital):
    rets, sharpes, dds = [], [], []
    for p in paths:
        m = stress_engine.run_single_backtest(p, cls, params, sim_kwargs, capital)
        rets.append(m.get("total_return_pct", 0.0))
        sharpes.append(m.get("sharpe_ratio", 0.0))
        dds.append(m.get("max_drawdown_pct", 0.0))
    a = np.array(rets)
    return {
        "n_paths": len(paths),
        "return_p5": round(float(np.percentile(a, 5)), 2),
        "return_p50": round(float(np.percentile(a, 50)), 2),
        "return_p95": round(float(np.percentile(a, 95)), 2),
        "mean_return": round(float(a.mean()), 2),
        "prob_profit_pct": round(float((a > 0).mean() * 100), 1),
        "mean_sharpe": round(float(np.mean(sharpes)), 3),
        "mean_max_dd": round(float(np.mean(dds)), 2),
    }


def section_ai_forecast(fetcher, defects, quick=False):
    print("\n=== SECTION 3: AI forecast (forward / crisis / paper) ===")
    rows = []
    combos = ([("BTC/USDT", "binance", "USD", "DCA"),
               ("BTC/USDT", "binance", "USD", "RSI"),
               ("RELIANCE", "nse", "INR", "PLA")]
              if not quick else [("BTC/USDT", "binance", "USD", "DCA")])
    start, end = TF_WINDOWS["1d"]
    horizon = 90
    n_paths = 50 if not quick else 20

    for symbol, source, currency, strat in combos:
        capital = CAPITAL_INR if currency == "INR" else CAPITAL_USD
        try:
            df = fetcher.fetch(symbol, start, end, source=source, interval="1d")
        except Exception as exc:
            defects.append(("AI_FETCH", symbol, str(exc)))
            continue
        cls = STRATEGY_REGISTRY[strat]
        params = dict(cls.default_params())
        if strat == "GRID":
            _ensure_grid_bounds(params, df)
        use_indian = source in ("nse", "bse") or is_indian(symbol)
        sim_kwargs = dict(symbol=symbol, use_indian_costs=use_indian,
                          market_type="equity_delivery" if use_indian else "crypto",
                          lot_size=1)

        # 3a. Forward test — pure block-bootstrap forward paths
        try:
            paths = forecast_engine.generate_paths(df, n_paths, horizon, seed=7)
            agg = _aggregate_paths(df, paths, cls, params, sim_kwargs, capital)
            rows.append({"mode": "Forward Test", "symbol": symbol, "strategy": strat,
                         "horizon_bars": horizon, **agg,
                         "scenario": "-", "status": "PASS"})
            print(f"  Forward  {symbol:10} {strat:5} P50={agg['return_p50']:+.1f}% "
                  f"P(profit)={agg['prob_profit_pct']:.0f}%")
        except Exception as exc:
            defects.append(("AI_FORWARD", f"{strat} on {symbol}", f"{type(exc).__name__}: {exc}"))
            rows.append({"mode": "Forward Test", "symbol": symbol, "strategy": strat,
                         "status": "ERROR", "error": f"{type(exc).__name__}: {exc}"})

        # 3b. Crisis overlay — forward paths with a stress scaffold applied
        for scen in (["luna_collapse", "covid_crash"] if not quick else ["luna_collapse"]):
            try:
                paths = forecast_engine.generate_paths(df, n_paths, horizon, seed=11)
                scen_obj = stress_engine.SCENARIO_PRESETS[scen]
                stressed = [stress_engine.apply_stress(p, scen_obj, 1.0, seed=i)
                            for i, p in enumerate(paths)]
                agg = _aggregate_paths(df, stressed, cls, params, sim_kwargs, capital)
                rows.append({"mode": "Crisis Sim", "symbol": symbol, "strategy": strat,
                             "horizon_bars": horizon, "scenario": scen, **agg,
                             "status": "PASS"})
                print(f"  Crisis   {symbol:10} {strat:5} [{scen}] "
                      f"P50={agg['return_p50']:+.1f}%")
            except Exception as exc:
                defects.append(("AI_CRISIS", f"{strat}/{scen} on {symbol}",
                                f"{type(exc).__name__}: {exc}"))
                rows.append({"mode": "Crisis Sim", "symbol": symbol, "strategy": strat,
                             "scenario": scen, "status": "ERROR",
                             "error": f"{type(exc).__name__}: {exc}"})

        # 3c. Paper trade — single forward path, run the strategy bar-by-bar
        try:
            path = forecast_engine.generate_one_path(df, horizon, seed=21)
            m = stress_engine.run_single_backtest(path, cls, params, sim_kwargs, capital)
            rows.append({"mode": "Paper Trade", "symbol": symbol, "strategy": strat,
                         "horizon_bars": horizon, "scenario": "-", "n_paths": 1,
                         "return_p50": round(m.get("total_return_pct", 0), 2),
                         "mean_return": round(m.get("total_return_pct", 0), 2),
                         "mean_sharpe": round(m.get("sharpe_ratio", 0), 3),
                         "mean_max_dd": round(m.get("max_drawdown_pct", 0), 2),
                         "prob_profit_pct": 100.0 if m.get("total_return_pct", 0) > 0 else 0.0,
                         "status": "PASS"})
            print(f"  Paper    {symbol:10} {strat:5} "
                  f"return={m.get('total_return_pct', 0):+.1f}% trades={m.get('num_trades', 0)}")
        except Exception as exc:
            defects.append(("AI_PAPER", f"{strat} on {symbol}", f"{type(exc).__name__}: {exc}"))
            rows.append({"mode": "Paper Trade", "symbol": symbol, "strategy": strat,
                         "status": "ERROR", "error": f"{type(exc).__name__}: {exc}"})
    return rows


# ─────────────────────────────────────────────────────────────────────────────
# Section 4 — Indicator engine validation
# ─────────────────────────────────────────────────────────────────────────────

def section_indicators(fetcher, defects):
    print("\n=== SECTION 4: Indicator engine validation ===")
    rows = []
    start, end = TF_WINDOWS["1d"]
    df = fetcher.fetch("BTC/USDT", start, end, source="binance", interval="1d")
    for ind in INDICATOR_CATALOG:
        key = ind["key"]
        outputs = ind.get("outputs", [])
        try:
            res = compute(df, key)
            cols = list(res.columns) if isinstance(res, pd.DataFrame) else [getattr(res, "name", key)]
            missing = [o for o in outputs if o not in cols]
            nan_frac = float(pd.DataFrame(res).isna().mean().mean()) if len(res) else 1.0
            status = "PASS" if not missing else "FAIL"
            if missing:
                defects.append(("INDICATOR", key, f"missing outputs {missing}"))
            rows.append({
                "key": key, "label": ind.get("label", key), "group": ind.get("group", ""),
                "expected_outputs": ", ".join(outputs),
                "produced_outputs": ", ".join(cols),
                "missing": ", ".join(missing) if missing else "-",
                "nan_fraction": round(nan_frac, 3),
                "rows": len(res), "status": status,
            })
        except Exception as exc:
            defects.append(("INDICATOR", key, f"{type(exc).__name__}: {exc}"))
            rows.append({"key": key, "label": ind.get("label", key),
                         "group": ind.get("group", ""), "status": "ERROR",
                         "error": f"{type(exc).__name__}: {exc}"})
    n_pass = sum(1 for r in rows if r["status"] == "PASS")
    print(f"  {n_pass}/{len(rows)} indicators OK")
    return rows


# ─────────────────────────────────────────────────────────────────────────────
# Excel writing helpers
# ─────────────────────────────────────────────────────────────────────────────

HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(bold=True, size=14, color="111827")
SUB_FONT = Font(italic=True, size=9, color="6B7280")
THIN = Side(style="thin", color="E5E7EB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
GREEN = PatternFill("solid", fgColor="DCFCE7")
RED = PatternFill("solid", fgColor="FEE2E2")
YELLOW = PatternFill("solid", fgColor="FEF9C3")
GREY = PatternFill("solid", fgColor="F3F4F6")


def _autosize(ws, max_w=46):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        longest = max((len(str(c.value)) for c in col if c.value is not None), default=8)
        ws.column_dimensions[letter].width = min(max(longest + 2, 10), max_w)


def write_sheet(ws, title, subtitle, df: pd.DataFrame, color_rules=None):
    ws.sheet_view.showGridLines = False
    ws["A1"] = title; ws["A1"].font = TITLE_FONT
    ws["A2"] = subtitle; ws["A2"].font = SUB_FONT
    start_row = 4
    cols = list(df.columns)
    for j, c in enumerate(cols, 1):
        cell = ws.cell(start_row, j, c.replace("_", " ").title())
        cell.fill = HEADER_FILL; cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    for i, (_, row) in enumerate(df.iterrows(), start_row + 1):
        for j, c in enumerate(cols, 1):
            v = row[c]
            if isinstance(v, float) and pd.isna(v):
                v = ""
            cell = ws.cell(i, j, v)
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if color_rules:
                color_rules(cell, c, row)
    ws.freeze_panes = ws.cell(start_row + 1, 1)
    _autosize(ws)


def status_colors(cell, col, row):
    if col == "status":
        cell.fill = {"PASS": GREEN, "ERROR": RED, "FAIL": RED}.get(str(row.get("status")), GREY)
    if col == "return_pct" and isinstance(row.get("return_pct"), (int, float)):
        cell.fill = GREEN if row["return_pct"] > 0 else (RED if row["return_pct"] < 0 else GREY)
    if col == "verdict":
        cell.fill = {"SURVIVED": GREEN, "DEGRADED": YELLOW, "BROKEN": RED,
                     "ERROR": RED}.get(str(row.get("verdict")), GREY)
    if col == "delta_pct" and isinstance(row.get("delta_pct"), (int, float)):
        cell.fill = GREEN if row["delta_pct"] >= -2 else (YELLOW if row["delta_pct"] >= -10 else RED)


# ─────────────────────────────────────────────────────────────────────────────
# Report builders
# ─────────────────────────────────────────────────────────────────────────────

def build_strategy_report(rows, outdir, stamp):
    df = pd.DataFrame(rows)
    wb = openpyxl.Workbook()
    # Leaderboard (passing only, by Sharpe)
    ok = df[df.status == "PASS"].copy()
    ws = wb.active; ws.title = "Leaderboard"
    lead = ok.sort_values("sharpe", ascending=False).head(60)
    write_sheet(ws, "Strategy Leaderboard — Top 60 by Sharpe",
                f"All strategies × assets × timeframes • generated {stamp}",
                lead[["asset_class", "symbol", "interval", "strategy", "category",
                      "num_trades", "return_pct", "sharpe", "sortino", "calmar",
                      "max_dd_pct", "win_rate", "regime_mix"]], status_colors)
    # Per asset-class sheets
    for ac in df.asset_class.dropna().unique():
        sub = df[df.asset_class == ac].sort_values(["symbol", "interval", "strategy"])
        ws = wb.create_sheet(ac[:28])
        write_sheet(ws, f"{ac} — full results",
                    f"{len(sub)} runs • generated {stamp}",
                    sub[["symbol", "interval", "strategy", "category", "candles",
                         "buy_signals", "num_trades", "return_pct", "ann_return_pct",
                         "sharpe", "sortino", "max_dd_pct", "win_rate",
                         "profit_factor", "final_equity", "status", "error"]],
                    status_colors)
    # Best strategy per asset
    if len(ok):
        idx = ok.groupby(["asset_class", "symbol", "interval"])["sharpe"].idxmax()
        best = ok.loc[idx].sort_values(["asset_class", "symbol"])
        ws = wb.create_sheet("Best Per Asset")
        write_sheet(ws, "Best strategy per asset / timeframe (by Sharpe)",
                    f"generated {stamp}",
                    best[["asset_class", "symbol", "interval", "strategy", "return_pct",
                          "sharpe", "sortino", "max_dd_pct", "win_rate", "num_trades"]],
                    status_colors)
    path = outdir / "Strategy_Backtest_Matrix.xlsx"
    wb.save(path); return path, df


def build_stress_report(rows, outdir, stamp):
    df = pd.DataFrame(rows)
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "All Runs"
    cols = [c for c in ["symbol", "currency", "strategy", "scenario", "severity",
                        "baseline_return_pct", "stressed_return_pct", "delta_pct",
                        "stressed_sharpe", "stressed_max_dd_pct", "mc_return_p50",
                        "mc_return_p5", "verdict", "status", "error"] if c in df.columns]
    write_sheet(ws, "Stress Test — all runs",
                f"{len(df)} runs • 17 scenarios × severities • generated {stamp}",
                df[cols], status_colors)
    # Scenario summary (median delta)
    ok = df[df.status == "PASS"].copy()
    if len(ok):
        summ = (ok.groupby("scenario")
                  .agg(median_delta=("delta_pct", "median"),
                       worst_delta=("delta_pct", "min"),
                       runs=("delta_pct", "size"))
                  .reset_index().sort_values("median_delta"))
        summ["verdict"] = summ["median_delta"].apply(
            lambda d: "SURVIVED" if d >= -2 else ("DEGRADED" if d >= -10 else "BROKEN"))
        summ = summ.round(2)
        ws = wb.create_sheet("Scenario Summary")
        write_sheet(ws, "Scenario impact summary (median Δ% across strategies/assets)",
                    f"generated {stamp}", summ, status_colors)
    path = outdir / "Stress_Test_Matrix.xlsx"
    wb.save(path); return path, df


def build_ai_report(rows, outdir, stamp):
    df = pd.DataFrame(rows)
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Forecast Runs"
    cols = [c for c in ["mode", "symbol", "strategy", "scenario", "horizon_bars",
                        "n_paths", "return_p5", "return_p50", "return_p95",
                        "mean_return", "prob_profit_pct", "mean_sharpe",
                        "mean_max_dd", "status", "error"] if c in df.columns]
    write_sheet(ws, "AI Forecast — forward / crisis / paper",
                f"Block-bootstrap pipeline (Kronos-ready) • {len(df)} runs • {stamp}",
                df[cols], status_colors)
    path = outdir / "AI_Forecast_Report.xlsx"
    wb.save(path); return path, df


def build_indicator_report(rows, outdir, stamp):
    df = pd.DataFrame(rows)
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Indicators"
    cols = [c for c in ["key", "label", "group", "expected_outputs",
                        "produced_outputs", "missing", "nan_fraction", "rows",
                        "status", "error"] if c in df.columns]
    write_sheet(ws, "Indicator Engine Validation",
                f"{len(df)} indicators in INDICATOR_CATALOG • {stamp}",
                df[cols], status_colors)
    path = outdir / "Indicator_Validation.xlsx"
    wb.save(path); return path, df


def build_defect_log(defects, outdir, stamp):
    df = pd.DataFrame(defects, columns=["category", "where", "detail"]) if defects \
        else pd.DataFrame([{"category": "-", "where": "-", "detail": "No defects found"}])
    df.insert(0, "id", [f"D{i:03d}" for i in range(1, len(df) + 1)])
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Defects"
    write_sheet(ws, "Defect Log", f"{len(defects)} issues • generated {stamp}", df)
    path = outdir / "Defect_Log.xlsx"
    wb.save(path); return path


def build_execution_summary(sections, defects, outdir, stamp, elapsed):
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Summary"
    ws.sheet_view.showGridLines = False
    ws["A1"] = "TradeVed Backtester — Comprehensive Feature Test"; ws["A1"].font = TITLE_FONT
    ws["A2"] = f"Generated {stamp} • runtime {elapsed:.0f}s • all features post stress-tester"
    ws["A2"].font = SUB_FONT

    summary_rows = []
    for name, df in sections.items():
        if df is None or not len(df):
            summary_rows.append({"section": name, "total": 0, "pass": 0,
                                 "error": 0, "pass_rate": "-"}); continue
        total = len(df)
        npass = int((df.get("status") == "PASS").sum()) if "status" in df else total
        nerr = total - npass
        summary_rows.append({"section": name, "total": total, "pass": npass,
                             "error": nerr,
                             "pass_rate": f"{100*npass/total:.1f}%" if total else "-"})
    sdf = pd.DataFrame(summary_rows)

    # Header
    r0 = 4
    heads = ["Section", "Total", "Pass", "Error", "Pass Rate"]
    for j, h in enumerate(heads, 1):
        c = ws.cell(r0, j, h); c.fill = HEADER_FILL; c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center"); c.border = BORDER
    for i, row in enumerate(summary_rows, r0 + 1):
        for j, key in enumerate(["section", "total", "pass", "error", "pass_rate"], 1):
            c = ws.cell(i, j, row[key]); c.border = BORDER
            c.alignment = Alignment(horizontal="center")
        # color pass-rate
        pr = ws.cell(i, 5)
        try:
            v = float(str(row["pass_rate"]).rstrip("%"))
            pr.fill = GREEN if v >= 99 else (YELLOW if v >= 90 else RED)
        except Exception:
            pass

    tot = int(sdf["total"].sum()); tp = int(sdf["pass"].sum())
    verdict_row = r0 + len(summary_rows) + 2
    ws.cell(verdict_row, 1, "TOTAL").font = Font(bold=True)
    ws.cell(verdict_row, 2, tot).font = Font(bold=True)
    ws.cell(verdict_row, 3, tp).font = Font(bold=True)
    ws.cell(verdict_row, 4, tot - tp).font = Font(bold=True)
    ws.cell(verdict_row, 5, f"{100*tp/tot:.1f}%" if tot else "-").font = Font(bold=True)

    gate = verdict_row + 2
    overall = 100 * tp / tot if tot else 0
    verdict = "GO" if overall >= 95 and len(defects) < 10 else \
              ("CONDITIONAL GO" if overall >= 85 else "NO-GO")
    ws.cell(gate, 1, "VERDICT").font = Font(bold=True, size=12)
    vcell = ws.cell(gate, 2, verdict)
    vcell.font = Font(bold=True, size=12,
                      color="065F46" if verdict == "GO" else "92400E" if "CONDITIONAL" in verdict else "991B1B")
    ws.cell(gate, 4, f"{len(defects)} defects logged")

    # Environment block
    env = gate + 3
    ws.cell(env, 1, "Environment").font = Font(bold=True)
    info = [
        ("Strategies registered", str(len(ALL_STRATEGIES))),
        ("Indicators in catalog", str(len(INDICATOR_CATALOG))),
        ("Stress scenarios", str(len(stress_engine.SCENARIO_PRESETS))),
        ("Asset classes", "Crypto, US Stock, Indian Equity"),
        ("Timeframes", ", ".join(TF_WINDOWS.keys())),
        ("Kronos GPU", "not connected (block-bootstrap fallback active)"),
        ("Python", sys.version.split()[0]),
    ]
    for k, (label, val) in enumerate(info, env + 1):
        ws.cell(k, 1, label); ws.cell(k, 2, val)
    _autosize(ws)
    path = outdir / "Execution_Summary.xlsx"
    wb.save(path); return path


# ─────────────────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────────────────

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--quick", action="store_true", help="smaller universe (smoke test)")
    ap.add_argument("--skip-ai", action="store_true")
    ap.add_argument("--skip-stress", action="store_true")
    args = ap.parse_args()

    stamp = datetime.now().strftime("%Y-%m-%d %H:%M")
    t_start = time.time()
    outdir = HERE / "comprehensive_reports"
    outdir.mkdir(exist_ok=True)

    fetcher = DataFetcher()
    validator = DataValidator()
    defects = []

    if args.quick:
        universe = QUICK_CRYPTO + QUICK_US + QUICK_INDIA
        timeframes = ["1d", "4h"]
    else:
        universe = CRYPTO + US + INDIA
        timeframes = ["1d", "4h", "1h", "15m"]

    print(f"TradeVed Comprehensive Test — {stamp}")
    print(f"  strategies={len(ALL_STRATEGIES)} assets={len(universe)} "
          f"timeframes={timeframes} quick={args.quick}")

    strat_rows = section_strategy_matrix(fetcher, validator, universe, timeframes, defects)
    stress_rows = [] if args.skip_stress else section_stress_matrix(fetcher, validator, defects, args.quick)
    ai_rows = [] if args.skip_ai else section_ai_forecast(fetcher, defects, args.quick)
    ind_rows = section_indicators(fetcher, defects)

    print("\n=== Writing Excel reports ===")
    p1, strat_df = build_strategy_report(strat_rows, outdir, stamp)
    print(f"  {p1.name}")
    stress_df = ind_df = ai_df = None
    if stress_rows:
        p2, stress_df = build_stress_report(stress_rows, outdir, stamp); print(f"  {p2.name}")
    if ai_rows:
        p3, ai_df = build_ai_report(ai_rows, outdir, stamp); print(f"  {p3.name}")
    p4, ind_df = build_indicator_report(ind_rows, outdir, stamp); print(f"  {p4.name}")
    pd_path = build_defect_log(defects, outdir, stamp); print(f"  {pd_path.name}")

    sections = {
        "Strategy Matrix": strat_df,
        "Stress Matrix": stress_df,
        "AI Forecast": ai_df,
        "Indicator Validation": ind_df,
    }
    ps = build_execution_summary(sections, defects, outdir, stamp, time.time() - t_start)
    print(f"  {ps.name}")

    # raw CSV dump
    strat_df.to_csv(outdir / "raw_results.csv", index=False)

    elapsed = time.time() - t_start
    total = sum(len(d) for d in sections.values() if d is not None)
    print(f"\nDONE in {elapsed:.0f}s • {total} total test rows • "
          f"{len(defects)} defects • reports in {outdir}")


if __name__ == "__main__":
    main()
